from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


SHEETS = [
    "00_Instrucciones",
    "01_Ficha_Proyecto",
    "02_Registro_Documentos",
    "03_Catalogo_Conductores",
    "04_Catalogo_Ductos",
    "05_Cables_Equipos",
    "06_Resumen_CAG",
    "07_Tramos_Camaras",
    "08_Matriz_Paso",
    "09_Resumen_PVC",
    "10_Observaciones",
    "11_Referencia_Documental",
]


def _style_header(ws: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"


def generate_workbook(output: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)

    wb["00_Instrucciones"].append(["Uso", "Completar solo datos confirmados o marcar PENDIENTE DE VALIDAR."])
    wb["01_Ficha_Proyecto"].append(["Campo", "Valor", "Estado"])
    for field in ["Proyecto", "Subestacion", "Cliente", "Revision", "Responsable"]:
        wb["01_Ficha_Proyecto"].append([field, "PENDIENTE DE VALIDAR", "PENDIENTE"])

    wb["02_Registro_Documentos"].append(
        ["Proyecto", "Subestacion", "Codigo documental", "Titulo", "Revision", "Fecha", "Tipo", "Fuente", "Estado"]
    )
    wb["03_Catalogo_Conductores"].append(["Conductor", "Diametro exterior mm", "Fuente", "Estado"])
    wb["04_Catalogo_Ductos"].append(["Ducto", "Diametro exterior mm", "Espesor mm", "Area util mm2", "Fuente", "Estado"])

    cables = wb["05_Cables_Equipos"]
    cables.append(
        [
            "ID cable",
            "Proyecto",
            "Pano",
            "Equipo origen",
            "Caja agrupamiento",
            "Camara ingreso",
            "Destino",
            "Funcion",
            "Conductor",
            "Cantidad",
            "Diametro exterior catalogo",
            "Diametro exterior manual",
            "Diametro utilizado",
            "Area unitaria",
            "Area total",
            "Clasificacion",
            "Fuente documental",
            "Pagina o lamina",
            "Estado",
            "Observacion",
        ]
    )
    for row in range(2, 102):
        cables[f"M{row}"] = f'=IF(ISNUMBER(L{row}),L{row},IF(ISNUMBER(K{row}),K{row},""))'
        cables[f"N{row}"] = f'=IF(ISNUMBER(M{row}),PI()*(M{row}/2)^2,"")'
        cables[f"O{row}"] = f'=IF(AND(ISNUMBER(J{row}),ISNUMBER(N{row})),J{row}*N{row},"")'

    wb["06_Resumen_CAG"].append(["Tramo", "Area total", "Ductos", "% ocupacion", "Estado"])
    wb["07_Tramos_Camaras"].append(["ID tramo", "Origen", "Destino", "Tipo ducto", "Cantidad ductos", "Fuente", "Estado"])
    matriz = wb["08_Matriz_Paso"]
    matriz.append(["ID cable", "Equipo", "Clasificacion", "Area total", "TR-01", "TR-02", "TR-03"])
    resumen = wb["09_Resumen_PVC"]
    resumen.append(
        [
            "Tramo",
            "Area control",
            "Ductos control",
            "% ocupacion control",
            "Area fuerza",
            "Ductos fuerza",
            "% ocupacion fuerza",
            "Reserva",
            "Total ductos",
            "Estado",
        ]
    )
    for row in range(2, 52):
        resumen[f"I{row}"] = f'=IF(AND(ISNUMBER(C{row}),ISNUMBER(F{row})),C{row}+F{row},"")'
    wb["10_Observaciones"].append(["ID", "Documento", "Pagina o lamina", "Observacion", "Estado", "Accion recomendada"])
    wb["11_Referencia_Documental"].append(["ID dato", "Documento", "Codigo", "Revision", "Pagina o lamina", "Nivel confianza"])

    for ws in wb.worksheets:
        _style_header(ws)
        for column_cells in ws.columns:
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(column_cells[0].value or "")) + 4, 14), 32)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Generar plantilla Excel para auditoria de canalizaciones.")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(generate_workbook(args.output))


if __name__ == "__main__":
    main()
