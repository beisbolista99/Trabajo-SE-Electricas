from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook

from common import DOCUMENT_RECORD_HEADERS, PENDING, backup_file, ensure_project_dir, infer_document_type, project_name_from_ficha

INPUT_DIRS = [
    "00_entrada/originales",
    "00_entrada/planos",
    "00_entrada/memorias_calculo",
    "00_entrada/especificaciones_tecnicas",
    "00_entrada/hctg",
    "00_entrada/catalogos",
    "00_entrada/comentarios_cliente",
    "00_entrada/otros",
]


def _load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"
    ws.append(DOCUMENT_RECORD_HEADERS)
    return wb


def register_documents(project: Path, output: Path | None = None) -> Path:
    project = ensure_project_dir(project)
    output = output or project / "registro_documentos.xlsx"
    wb = _load_or_create_workbook(output)
    ws = wb["Registro"]
    if ws.max_row == 0:
        ws.append(DOCUMENT_RECORD_HEADERS)

    existing_files = {str(row[12].value) for row in ws.iter_rows(min_row=2) if row[12].value}
    project_name = project_name_from_ficha(project)
    added = 0

    for input_dir in INPUT_DIRS:
        folder = project / input_dir
        if not folder.exists():
            continue
        for file in sorted(folder.rglob("*")):
            if not file.is_file() or file.name.startswith("~$") or file.name == ".gitkeep":
                continue
            rel = file.relative_to(project).as_posix()
            if rel in existing_files:
                continue
            doc_type = infer_document_type(file)
            ws.append(
                [
                    project_name,
                    PENDING,
                    PENDING,
                    file.stem,
                    PENDING,
                    PENDING,
                    PENDING,
                    doc_type,
                    PENDING,
                    PENDING,
                    rel,
                    "PENDIENTE",
                    rel,
                    "Metadata inicial; validar contra contenido interno.",
                ]
            )
            added += 1

    if output.exists() and added:
        backup_file(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Registrar documentos de entrada en Excel.")
    parser.add_argument("--project", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    print(register_documents(args.project, args.output))


if __name__ == "__main__":
    main()

